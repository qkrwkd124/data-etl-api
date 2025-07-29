"""EIU 주요 경제지표 데이터 ETL 처리 서비스.

EIU (Economist Intelligence Unit) 주요 경제지표 데이터의 추출, 변환, 적재를 담당합니다.
복잡한 엑셀 구조 파싱과 데이터 정규화를 포함합니다.
"""

import pandas as pd
import asyncio
from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
import traceback
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.eiu_repository import EIUEconomicIndicatorRepository
from app.repositories.history_repository import DataUploadAutoHistoryRepository
from app.schemas.eiu_schemas import ExcelRowData
from app.models.EIU import EconomicData
from app.core.constants.eiu import EIU_CODES, EIU_COLUMN_MAPPING, EIU_ESTIMATE_COLOR, EIUDataType
from app.core.constants.error import ErrorMessages, ErrorCode
from app.core.exceptions import (
    DataProcessingException,
    DatabaseException,
    FileException,
    DataNotFoundException,
    ValidationException
)
from app.core.logger import get_logger
from app.utils.file_utils import validate_file


logger = get_logger()

def _get_cell_color(cell) -> Optional[str]:
    """엑셀 셀의 배경색 RGB 값 추출.

    openpyxl 셀 객체에서 배경색을 추출하여
    예측 데이터 식별에 사용합니다.
    
    Args:
        cell: openpyxl 셀 객체
        
    Returns:
        Optional[str]: RGB 문자열 또는 None (배경색이 없는 경우)
    """
    if cell.fill.patternType == 'solid':
        return cell.font.color.rgb
    return None

def _find_header_row(sheet) -> Optional[int]:
    """엑셀 시트에서 헤더 행 찾기.
    
    'Series'와 'Code' 컬럼이 있는 행을 찾아
    데이터 시작 위치를 파악합니다.
    
    Args:
        sheet: openpyxl 워크시트 객체
        
    Returns:
        Optional[int]: 헤더 행 번호 (1부터 시작) 또는 None
    """
    for row_idx in range(1, 20): # 상위 20행 내에서 헤더 찾기기
        series_cell = sheet.cell(row=row_idx, column=1).value
        code_cell = sheet.cell(row=row_idx, column=2).value
        
        if series_cell == "Series" and code_cell == "Code":
            return row_idx
        
    return None


def _extract_column_names(sheet, header_row: int) -> List[str]:
    """헤더 행에서 컬럼명 추출.
    
    지정된 헤더 행에서 모든 컬럼명을 추출하여
    리스트로 반환합니다.
    
    Args:
        sheet: openpyxl 워크시트 객체
        header_row (int): 헤더 행 번호
        
    Returns:
        List[str]: 컬럼명 리스트
    """
    column_names = []
    col_idx = 1
    while True:
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if cell_value is None:
            break
        column_names.append(cell_value)
        col_idx += 1
    return column_names

def _create_excel_row_from_sheet_data(
    sheet, 
    row_idx: int, 
    column_names: List[str], 
    country_code: str
) -> Optional[ExcelRowData]:
    """시트 데이터에서 ExcelRowData 객체 생성.
    
    엑셀 행 데이터를 읽어서 구조화된 객체로 변환합니다.
    연도별 데이터와 셀 색상 정보를 포함하여 처리합니다.
    
    Args:
        sheet: openpyxl 워크시트 객체
        row_idx (int): 처리할 행 번호
        column_names (List[str]): 컬럼명 리스트
        country_code (str): 국가 코드
        
    Returns:
        Optional[ExcelRowData]: 변환된 행 데이터 객체 또는 None
    """
    row_data = {
        "country_code": country_code,
        "year_data": {}
    }

    for col_idx, col_name in enumerate(column_names, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell_value = str(cell.value)

        #일반 필드 처리
        if col_name in EIU_COLUMN_MAPPING:
            row_data[EIU_COLUMN_MAPPING[col_name]] = cell_value
            
        #연도 데이터 처리
        elif isinstance(col_name, str) and col_name.strip().isdigit():
            color = _get_cell_color(cell)
            if cell_value and cell_value != "–" :
                cell_value = round(float(cell_value), 1)
                if color == EIU_ESTIMATE_COLOR : # 블루
                    year_value = f"{EIUDataType.ESTIMATE.value}|{cell_value}"
                else :
                    year_value = f"{EIUDataType.ACTUAL.value}|{cell_value}"
            else :
                year_value = EIUDataType.MISSING.value
            
            row_data["year_data"][col_name] = year_value
    
    excel_row = ExcelRowData(**row_data)
    
    return excel_row


def _create_default_excel_row(code: str, country_code: str, series: str, year_columns: List[str]) -> ExcelRowData:
    """
    누락된 코드에 대한 기본 스키마 객체 생성
    """
    default_year_data = {year: EIUDataType.MISSING.value for year in year_columns}
    
    return ExcelRowData(
        country_code=country_code,
        code=code,
        series=series,
        currency="", 
        units="",
        source="",
        definition="",
        note="", 
        published="",
        year_data=default_year_data
    )


def _convert_to_dataframe(excel_rows: List[ExcelRowData],year_columns: List[str]) -> pd.DataFrame:
    """
    ExcelRowData 리스트를 DataFrame으로 변환
    """
    
    if not excel_rows :
        return pd.DataFrame()
    
    df_data = [excel_row.to_dataframe_dict(year_columns) for excel_row in excel_rows]
    
    df = pd.DataFrame(df_data)

    if year_columns :
        max_year = max(map(int,year_columns))
        max_YY = max_year % 100

        for i in range(max_YY + 1, 52) :
            df[f"eiu_year{i}"] = EIUDataType.FORECAST.value

    return df

async def process_data(file_path:str) :

    logger.info(f"원본 파일 처리 시작 : {file_path}")

    # 확장자 오류로 인한 바이너리 스트림으로 처리
    with open(file_path, "rb") as file:
        workbook = openpyxl.load_workbook(file)

    sheet_names = workbook.sheetnames

    # 모든 Excel 행 데이터
    all_excel_rows = []

    for sheet_name in sheet_names:
        logger.info(f"시트 처리중 : {sheet_name}")
        sheet = workbook[sheet_name]

        header_row = None

        # 헤더 행 찾기 (보통 'Series'와 'Code' 컬럼이 있는 행)
        header_row = _find_header_row(sheet)
        if header_row is None:
            logger.error(f"헤더 행을 찾을 수 없음: {sheet_name}")
            raise FileException(
                message=ErrorMessages.get_message(ErrorCode.FILE_HEADER_NOT_FOUND),
                error_code=ErrorCode.FILE_HEADER_NOT_FOUND,
                detail={
                    "file_path": file_path,
                    "sheet_name": sheet_name
                }
            )

        # 컬럼 이름 추출
        column_names = _extract_column_names(sheet, header_row)

        logger.info(f"열 이름 : {column_names}")

        #연도 컬럼 식별
        year_columns = [col for col in column_names if isinstance(col,str) and col.strip().isdigit()]

        # 시트의 데이터를 코드별로 저장할 딕셔너리
        sheet_data_by_code = {}

        #데이터 행 처리
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            
            # 스키마 객체 생성
            excel_row = _create_excel_row_from_sheet_data(
                sheet, row_idx, column_names, sheet_name
            )

            if excel_row and excel_row.code and excel_row.code in list(EIU_CODES.keys()) :
                sheet_data_by_code[excel_row.code] = excel_row


        # 누락된 코드에 대한 기본 데이터 생성
        for code, series in EIU_CODES.items() :
            if code not in sheet_data_by_code :
                default_row = _create_default_excel_row(code, sheet_name, series, year_columns)
                sheet_data_by_code[code] = default_row

        all_excel_rows.extend(sheet_data_by_code.values())

    df = _convert_to_dataframe(all_excel_rows, year_columns)

    return df

async def process_eiu_economic_indicator(
        seq: int,
        db: AsyncSession,
        replace_all: bool = True
) :
    """EIU 주요 경제지표 파일 전체 처리.
    
    파일 검증부터 데이터 가공, 데이터베이스 저장까지
    전체 ETL 프로세스를 수행합니다.
    
    Args:
        seq (int): 파일 처리 순번 (이력 조회용)
        dbprsr (AsyncSession): 메인 데이터베이스 세션
        dbpdtm (AsyncSession): 이력 데이터베이스 세션
        replace_all (bool): True면 전체 교체, False면 추가 (기본값: True)
        
    Returns:
        pd.DataFrame: 처리된 데이터프레임
        
    Raises:
        DataProcessingException: 데이터 처리 실패 시
        DatabaseException: 데이터베이스 작업 실패 시
        FileException: 파일 검증 또는 읽기 실패 시
        DataNotFoundException: 파일 정보 조회 실패 시
        ValidationException: 데이터 검증 실패 시
    """

    try :
        # Repository 객체 생성
        repository = EIUEconomicIndicatorRepository(db)
        history_repository = DataUploadAutoHistoryRepository(db)

        history_info = await history_repository.get_history_info(seq)
        file_path = str(Path(history_info.file_path_nm,history_info.file_nm))
        logger.info(f"EIU Economic Indicator 파일 처리 시작: {file_path}")

        # 0. 이력 데이터 삽입
        await history_repository.start_processing(seq)

        # 1. 파일 유효성 검사
        await validate_file(file_path, history_info.file_exts_nm)

        # 2. 데이터 가공
        df = await process_data(file_path)

        if df.empty :
            logger.error("처리할 수 있는 데이터가 없습니다.")
            raise DataProcessingException(
                message=ErrorMessages.get_message(ErrorCode.DATA_PROCESSING_ERROR),
                error_code=ErrorCode.DATA_PROCESSING_ERROR,
                detail={
                    "file_path": file_path,
                }
            )

        try :
            #2. 국가명 매핑
            country_mapping = await repository.get_country_mapping()
            df["eiu_cont_en_nm"] = df["eiu_country_code"].map(country_mapping).fillna('')

        except Exception as e:
            logger.error(f"국가명 매핑 중 오류: \n{traceback.format_exc()}")
            raise DataProcessingException(
                message=ErrorMessages.get_message(ErrorCode.DATA_PROCESSING_ERROR),
                error_code=ErrorCode.DATA_PROCESSING_ERROR,
                detail={
                    "file_path": file_path,
                }
            )
        
        try :
        #. 2. 데이터베이스 저장
            logger.info("2. 데이터베이스 저장 중...")

            if replace_all :
                db_result = await repository.replace_all_data(df)
            else :
                insert_count = await repository.insert_dataframe(df)
                await db.commit()

            logger.info(f"데이터베이스 저장 완료: {db_result}")

        except Exception as e:
            logger.error(f"데이터베이스 저장 중 오류: \n{traceback.format_exc()}")
            raise DatabaseException(
                message=ErrorMessages.get_message(ErrorCode.DATABASE_ERROR),
                error_code=ErrorCode.DATABASE_ERROR,
                detail={
                    "file_path": file_path,
                }
            )

        #. 3. CSV 저장

        logger.info(f"EIU Economic Indicator 파일 처리 완료: {file_path}")
        await history_repository.success_processing(seq,
                                                    result_table_name=EconomicData.__tablename__,
                                                    process_count=len(df),
                                                    message=ErrorMessages.SUCCESS)
        return df
    
    except (DataProcessingException, DatabaseException, FileException, DataNotFoundException, ValidationException) as e:
        logger.error(f"처리 실패: {e.error_code.value} - {e.message}")
        await history_repository.fail_processing(seq, message=e.message)
        raise 
    except Exception as e:
        logger.error(f"예상하지 못한 시스템 오류: \n{traceback.format_exc()}")
        await history_repository.fail_processing(seq, message=ErrorMessages.get_message(ErrorCode.SYSTEM_ERROR))
        raise e


if __name__ == "__main__" :

    # current_file_path = Path(os.getcwd()) 
    # if str(current_file_path) not in sys.path:
    #     sys.path.insert(0, str(current_file_path))

    file_path = "/appdata/storage/research/original/2.EIU_AllDataByGeography_로데이터.xlsx"

    data = asyncio.run(process_data(file_path))
    print( data[data["eiu_country_code"] == "CA"][['eiu_year20','eiu_year21','eiu_year22','eiu_year23','eiu_year24']] )
    
